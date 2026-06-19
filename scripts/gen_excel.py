"""
gen_excel.py — EuroAir Intel Manufacturer Report
Generates a professional Excel workbook with 5 sheets:
  1. Summary Dashboard
  2. Monthly Deliveries (with chart)
  3. Model Breakdown (with chart)
  4. Backlog Analysis (with chart)
  5. Stock Data
"""

import json
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, Reference, PieChart
    from openpyxl.chart.series import DataPoint
    from openpyxl.chart.label import DataLabel
except ImportError:
    os.system("pip install openpyxl --break-system-packages -q")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, Reference, PieChart
    from openpyxl.chart.series import DataPoint

with open("output/manufacturer_data.json") as f:
    d = json.load(f)

B = d["boeing"]
A = d["airbus"]
Bs = B.get("stock", {})
As = A.get("stock", {})

# ── Style helpers ──────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)

def font(bold=False, size=11, color="1E293B", italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic, name="Calibri")

def border_thin():
    s = Side(style="thin", color="E2E8F0")
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium():
    s = Side(style="medium", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# Color palette
NAVY    = "0D1F3C"
BOEING  = "0033A0"
AIRBUS  = "005B8E"
GOLD    = "C9A84C"
OFFWHITE= "F8FAFC"
LGRAY   = "E2E8F0"
SLATE   = "64748B"
WHITE   = "FFFFFF"
GREEN   = "16A34A"
RED     = "DC2626"
BLIGHT  = "D6E4FF"
ALIGHT  = "D0EEFF"

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def header_row(ws, row, values, colors=None, text_color=WHITE, size=11, bold=True):
    for i, val in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=val)
        c.font = font(bold=bold, size=size, color=text_color)
        c.fill = fill(colors[i-1] if colors and i <= len(colors) else NAVY)
        c.alignment = align("center")
        c.border = border_thin()

def data_row(ws, row, values, bg=WHITE, bold=False, aligns=None, colors=None):
    for i, val in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=val)
        c.font = font(bold=bold, size=10, color=colors[i-1] if colors else "1E293B")
        c.fill = fill(bg)
        c.alignment = align(aligns[i-1] if aligns else "left")
        c.border = border_thin()

def title_block(ws, row, title, subtitle, span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=title)
    c.font = font(bold=True, size=18, color=WHITE)
    c.fill = fill(NAVY)
    c.alignment = align("center")

    ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=span)
    c2 = ws.cell(row=row+1, column=1, value=subtitle)
    c2.font = font(bold=False, size=10, color=SLATE, italic=True)
    c2.fill = fill(OFFWHITE)
    c2.alignment = align("center")

# ── Build workbook ─────────────────────────────────────────────────────────────
wb = Workbook()

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1: Summary Dashboard
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Summary Dashboard"
ws1.sheet_view.showGridLines = False
ws1.row_dimensions[1].height = 40
ws1.row_dimensions[2].height = 22

title_block(ws1, 1, f"EuroAir Intel — Manufacturer Intelligence Report", f"{d['reportMonth']}  ·  Data as of {d['asOf']}  ·  Sources: Boeing.com, Airbus.com, Yahoo Finance")

# Column widths
for col, w in enumerate([22, 18, 18, 4, 22, 18, 18], 1):
    set_col_width(ws1, col, w)

# Section headers
ws1.row_dimensions[4].height = 26
for col, (val, fc) in enumerate(zip(["Metric", "Boeing", "Airbus", "", "Metric", "Boeing", "Airbus"],
                                     [NAVY, BOEING, AIRBUS, WHITE, NAVY, BOEING, AIRBUS]), 1):
    c = ws1.cell(row=4, column=col, value=val)
    c.font = font(bold=True, size=11, color=WHITE if fc != WHITE else SLATE)
    c.fill = fill(fc)
    c.alignment = align("center")

scorecard = [
    ("Deliveries YTD",   B["deliveries_ytd"],  A["deliveries_ytd"],  "Orders YTD",      B["orders_ytd"],     A["orders_ytd"]),
    ("Total Backlog",    B["backlog"],          A["backlog"],          "Backlog (years)", B["backlog_years"],  A["backlog_years"]),
    ("737 MAX Deliv.",   B["models"].get("737 MAX",0), A["models"].get("A320neo",0), "Wide-body Deliv.", B["models"].get("787",0)+B["models"].get("777",0), A["models"].get("A350",0)+A["models"].get("A330",0)),
    ("BA Stock Price",   f"${Bs.get('price','N/A')}", f"€{As.get('price','N/A')}", "Stock Chg Today", f"{Bs.get('change_pct',0):+.2f}%", f"{As.get('change_pct',0):+.2f}%"),
]

for i, (m1, b1, a1, m2, b2, a2) in enumerate(scorecard, 5):
    ws1.row_dimensions[i].height = 22
    bg = OFFWHITE if i % 2 == 0 else WHITE
    for col, (val, bc) in enumerate([(m1, OFFWHITE), (b1, BLIGHT), (a1, ALIGHT), ("", WHITE), (m2, OFFWHITE), (b2, BLIGHT), (a2, ALIGHT)], 1):
        c = ws1.cell(row=i, column=col, value=val)
        c.font = font(bold=col in [2, 3, 6, 7], size=11, color=BOEING if col == 2 else AIRBUS if col == 3 else BOEING if col == 6 else AIRBUS if col == 7 else "1E293B")
        c.fill = fill(bc)
        c.alignment = align("center")
        c.border = border_thin()

# Source note
ws1.row_dimensions[10].height = 18
ws1.merge_cells("A10:G10")
src = ws1.cell(row=10, column=1, value=f"Source: Boeing.com (official monthly XLS) · Airbus.com (official monthly XLS) · Yahoo Finance (yfinance). Generated {d['reportDate']}.")
src.font = font(size=8, color=SLATE, italic=True)
src.fill = fill(OFFWHITE)
src.alignment = align("left")

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2: Monthly Deliveries
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Monthly Deliveries")
ws2.sheet_view.showGridLines = False
set_col_width(ws2, 1, 14)
set_col_width(ws2, 2, 16)
set_col_width(ws2, 3, 16)
set_col_width(ws2, 4, 16)
set_col_width(ws2, 5, 16)

title_block(ws2, 1, "Monthly Deliveries 2026 YTD", f"Boeing vs Airbus  ·  {d['reportMonth']}", span=5)

header_row(ws2, 4, ["Month", "Boeing", "Airbus", "Combined", "Boeing Lead (+/-)"],
           colors=[NAVY, BOEING, AIRBUS, SLATE, NAVY])

months = d.get("chartLabels", [])
b_series = d.get("boeingMonthlySeries", [])
a_series = d.get("airbusMonthlySeries", [])

for i, (m, bv, av) in enumerate(zip(months, b_series, a_series), 5):
    ws2.row_dimensions[i].height = 20
    bg = OFFWHITE if i % 2 == 0 else WHITE
    combined = bv + av
    lead = bv - av
    data_row(ws2, i, [m, bv, av, combined, lead],
             bg=bg,
             aligns=["center","center","center","center","center"],
             colors=["1E293B", BOEING, AIRBUS, SLATE, GREEN if lead >= 0 else RED])

# Totals row
tr = len(months) + 5
ws2.row_dimensions[tr].height = 24
total_b = sum(b_series)
total_a = sum(a_series)
data_row(ws2, tr,
         ["TOTAL YTD", total_b, total_a, total_b+total_a, total_b-total_a],
         bg=NAVY, bold=True,
         aligns=["center","center","center","center","center"],
         colors=[WHITE, "90CAFF", "90E0FF", WHITE, "90FFB0" if total_b >= total_a else "FFB0B0"])

# Chart
chart = BarChart()
chart.type = "col"
chart.grouping = "clustered"
chart.title = f"Monthly Deliveries 2026 YTD"
chart.y_axis.title = "Aircraft Delivered"
chart.x_axis.title = "Month"
chart.style = 10
chart.width = 22
chart.height = 14
chart.shape = 4

data_ref = Reference(ws2, min_col=2, max_col=3, min_row=4, max_row=4+len(months))
cats_ref = Reference(ws2, min_col=1, min_row=5, max_row=4+len(months))
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.series[0].graphicalProperties.solidFill = BOEING
chart.series[1].graphicalProperties.solidFill = "00AEEF"
ws2.add_chart(chart, f"G4")

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3: Model Breakdown
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Model Breakdown")
ws3.sheet_view.showGridLines = False
for col, w in enumerate([20, 16, 16, 16], 1):
    set_col_width(ws3, col, w)

title_block(ws3, 1, "Deliveries by Aircraft Model — 2026 YTD", f"Boeing vs Airbus  ·  {d['reportMonth']}", span=4)

# Boeing models
ws3.row_dimensions[4].height = 22
header_row(ws3, 4, ["Boeing Model", "Deliveries YTD", "Backlog", "Backlog/Deliv Ratio"],
           colors=[BOEING, BOEING, BOEING, BOEING])
b_models = B.get("models", {})
b_backlog = B.get("backlog_by_model", {})
for i, (model, deliv) in enumerate(b_models.items(), 5):
    bl = b_backlog.get(model, 0)
    ratio = f"={get_column_letter(3)}{i}/{get_column_letter(2)}{i}" if deliv > 0 else "N/A"
    ws3.row_dimensions[i].height = 20
    bg = BLIGHT if i % 2 == 0 else WHITE
    data_row(ws3, i, [model, deliv, bl, bl/deliv if deliv > 0 else 0],
             bg=bg, aligns=["left","center","center","center"],
             colors=["1E293B", BOEING, NAVY, SLATE])
    ws3.cell(row=i, column=4).number_format = "0.0x"

# Spacer
ws3.row_dimensions[10].height = 14

# Airbus models
ws3.row_dimensions[11].height = 22
header_row(ws3, 11, ["Airbus Model", "Deliveries YTD", "Backlog", "Backlog/Deliv Ratio"],
           colors=[AIRBUS, AIRBUS, AIRBUS, AIRBUS])
a_models = A.get("models", {})
a_backlog = A.get("backlog_by_model", {})
for i, (model, deliv) in enumerate(a_models.items(), 12):
    bl = a_backlog.get(model, 0)
    ws3.row_dimensions[i].height = 20
    bg = ALIGHT if i % 2 == 0 else WHITE
    data_row(ws3, i, [model, deliv, bl, bl/deliv if deliv > 0 else 0],
             bg=bg, aligns=["left","center","center","center"],
             colors=["1E293B", AIRBUS, NAVY, SLATE])
    ws3.cell(row=i, column=4).number_format = "0.0x"

# Pie chart — Boeing model share
pie_b = PieChart()
pie_b.title = "Boeing Delivery Mix"
pie_b.style = 10
pie_b.width = 14
pie_b.height = 10
b_vals_ref   = Reference(ws3, min_col=2, min_row=5, max_row=5+len(b_models)-1)
b_labels_ref = Reference(ws3, min_col=1, min_row=5, max_row=5+len(b_models)-1)
pie_b.add_data(b_vals_ref)
pie_b.set_categories(b_labels_ref)
ws3.add_chart(pie_b, "F4")

# Pie chart — Airbus model share
pie_a = PieChart()
pie_a.title = "Airbus Delivery Mix"
pie_a.style = 10
pie_a.width = 14
pie_a.height = 10
a_vals_ref   = Reference(ws3, min_col=2, min_row=12, max_row=12+len(a_models)-1)
a_labels_ref = Reference(ws3, min_col=1, min_row=12, max_row=12+len(a_models)-1)
pie_a.add_data(a_vals_ref)
pie_a.set_categories(a_labels_ref)
ws3.add_chart(pie_a, "F20")

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4: Backlog Analysis
# ═══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Backlog Analysis")
ws4.sheet_view.showGridLines = False
for col, w in enumerate([22, 16, 16, 16, 16], 1):
    set_col_width(ws4, col, w)

title_block(ws4, 1, "Order Backlog Analysis", f"Boeing vs Airbus  ·  {d['reportMonth']}", span=5)

header_row(ws4, 4, ["Programme", "Boeing Backlog", "Airbus Backlog", "Boeing %", "Airbus %"],
           colors=[NAVY, BOEING, AIRBUS, BOEING, AIRBUS])

b_bl = B.get("backlog_by_model", {})
a_bl = A.get("backlog_by_model", {})
b_total_bl = B["backlog"]
a_total_bl = A["backlog"]

all_programmes = list(set(list(b_bl.keys()) + list(a_bl.keys())))
for i, prog in enumerate(all_programmes, 5):
    bv = b_bl.get(prog, 0)
    av = a_bl.get(prog, 0)
    ws4.row_dimensions[i].height = 20
    bg = OFFWHITE if i % 2 == 0 else WHITE
    b_pct = bv / b_total_bl if b_total_bl > 0 else 0
    a_pct = av / a_total_bl if a_total_bl > 0 else 0
    data_row(ws4, i, [prog, bv if bv else "—", av if av else "—", b_pct if bv else "—", a_pct if av else "—"],
             bg=bg, aligns=["left","center","center","center","center"],
             colors=["1E293B", BOEING, AIRBUS, BOEING, AIRBUS])
    if bv: ws4.cell(row=i, column=4).number_format = "0.0%"
    if av: ws4.cell(row=i, column=5).number_format = "0.0%"

# Totals
tr = len(all_programmes) + 5
ws4.row_dimensions[tr].height = 24
data_row(ws4, tr, ["TOTAL BACKLOG", b_total_bl, a_total_bl, "100%", "100%"],
         bg=NAVY, bold=True, aligns=["left","center","center","center","center"],
         colors=[WHITE, "90CAFF", "90E0FF", WHITE, WHITE])

# Runway row
rr = tr + 1
ws4.row_dimensions[rr].height = 22
data_row(ws4, rr, ["Production Runway (years)", B["backlog_years"], A["backlog_years"], "", ""],
         bg=GOLD, bold=True, aligns=["left","center","center","center","center"],
         colors=[NAVY, NAVY, NAVY, NAVY, NAVY])

# Bar chart comparing backlogs
chart4 = BarChart()
chart4.type = "bar"
chart4.grouping = "clustered"
chart4.title = "Backlog by Programme"
chart4.y_axis.title = "Aircraft on Order"
chart4.style = 10
chart4.width = 22
chart4.height = 14
data4  = Reference(ws4, min_col=2, max_col=3, min_row=4, max_row=4+len(all_programmes))
cats4  = Reference(ws4, min_col=1, min_row=5, max_row=4+len(all_programmes))
chart4.add_data(data4, titles_from_data=True)
chart4.set_categories(cats4)
chart4.series[0].graphicalProperties.solidFill = BOEING
chart4.series[1].graphicalProperties.solidFill = "00AEEF"
ws4.add_chart(chart4, "G4")

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5: Stock Data
# ═══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Stock Data")
ws5.sheet_view.showGridLines = False
for col, w in enumerate([22, 16, 16], 1):
    set_col_width(ws5, col, w)

title_block(ws5, 1, "Stock Performance — BA vs AIR.PA", f"As of {Bs.get('as_of', d['reportDate'])}  ·  Source: Yahoo Finance", span=3)

header_row(ws5, 4, ["Metric", "Boeing (BA)", "Airbus (AIR.PA)"],
           colors=[NAVY, BOEING, AIRBUS])

stock_rows = [
    ("Current Price",    f"${Bs.get('price','N/A')}",       f"€{As.get('price','N/A')}"),
    ("Daily Change $",   f"${Bs.get('change_abs',0):+.2f}", f"€{As.get('change_abs',0):+.2f}"),
    ("Daily Change %",   f"{Bs.get('change_pct',0):+.2f}%", f"{As.get('change_pct',0):+.2f}%"),
    ("52-Week High",     f"${Bs.get('week52_high','N/A')}", f"€{As.get('week52_high','N/A')}"),
    ("52-Week Low",      f"${Bs.get('week52_low','N/A')}",  f"€{As.get('week52_low','N/A')}"),
    ("Currency",         "USD (NYSE)",                        "EUR (Euronext Paris)"),
    ("Data Date",        Bs.get("as_of", d["reportDate"]),   As.get("as_of", d["reportDate"])),
]

for i, (metric, bval, aval) in enumerate(stock_rows, 5):
    ws5.row_dimensions[i].height = 22
    bg = OFFWHITE if i % 2 == 0 else WHITE
    data_row(ws5, i, [metric, bval, aval],
             bg=bg, aligns=["left","center","center"],
             colors=["1E293B", BOEING, AIRBUS])

# Source note
ws5.row_dimensions[13].height = 18
ws5.merge_cells("A13:C13")
src5 = ws5.cell(row=13, column=1, value="Source: Yahoo Finance via yfinance library. Prices reflect most recent trading day close. Not financial advice.")
src5.font = font(size=8, color=SLATE, italic=True)
src5.fill = fill(OFFWHITE)
src5.alignment = align("left")

# ── Save ───────────────────────────────────────────────────────────────────────
os.makedirs("output", exist_ok=True)
fname = f"output/Manufacturer_Data_{d['reportDate']}.xlsx"
wb.save(fname)
print(f"✓ Saved {fname}")
